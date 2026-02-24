import os
import json
import base64
import logging
import tempfile
from pathlib import Path
from pydantic import BaseModel, Field
from google import genai
from google.genai import types
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from telegram import Update
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    ContextTypes, filters,
)
from dotenv import load_dotenv
load_dotenv()

# ─── Logging ───
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ─── Config ───
TELEGRAM_BOT_TOKEN = os.getenv("telegram_token")
GOOGLE_API_KEY = os.getenv("gemini_api")



#  PYDANTIC MODELS

class ShiftDetail(BaseModel):
    shift_name: str = Field(default="")
    date: str = Field(default="")
    role: str = Field(default="")
    quantity: int = Field(default=0)
    start: str = Field(default="")
    finish: str = Field(default="")
    hours: float = Field(default=0.0)
    total_hours: float = Field(default=0.0)
    tariff: str = Field(default="")
    in_total: str = Field(default="")


class MiscItem(BaseModel):
    item: str = Field(default="")
    quantity: str = Field(default="")
    tariff: str = Field(default="")
    in_total: str = Field(default="")


class InvoiceData(BaseModel):
    supplier_name: str = Field(default="", alias="Supplier")
    supplier_address: str = Field(default="")
    supplier_registration_no: str = Field(default="")
    supplier_vat_number: str = Field(default="")
    client_name: str = Field(default="", alias="Client")
    client_address: str = Field(default="")
    client_registration_no: str = Field(default="")
    client_vat_number: str = Field(default="")
    order_name: str = Field(default="")
    order_number: str = Field(default="")
    date_range: str = Field(default="")
    location: str = Field(default="")
    shifts: list[ShiftDetail] = Field(default_factory=list)
    job_total_quantity: int = Field(default=0)
    total_hours: float = Field(default=0.0)
    job_total_amount: str = Field(default="")
    misc_items: list[MiscItem] = Field(default_factory=list)
    misc_total: str = Field(default="")
    total_price_excl_vat: str = Field(default="")
    vat_percentage: str = Field(default="")
    vat_amount: str = Field(default="")
    total_price_incl_vat: str = Field(default="")

    class Config:
        populate_by_name = True




#  GEMINI EXTRACTION

EXTRACTION_PROMPT = """You are a precise OCR data extraction engine.

Extract ALL data from this invoice/price quote image into the following JSON structure.
Be exact with numbers, currencies, dates, and formatting as they appear in the document.
If a field is not present in the document, use an empty string "" for text or 0 for numbers.

Return ONLY valid JSON — no markdown, no explanation:

{
  "Supplier": "company name",
  "supplier_address": "full address",
  "supplier_registration_no": "registration number",
  "supplier_vat_number": "VAT number",
  "Client": "company name",
  "client_address": "full address",
  "client_registration_no": "registration number",
  "client_vat_number": "VAT number",
  "order_name": "order/event name",
  "order_number": "order ID",
  "date_range": "start - end date",
  "location": "venue address",
  "shifts": [
    {
      "shift_name": "shift identifier",
      "date": "date",
      "role": "worker role",
      "quantity": 1,
      "start": "HH:MM",
      "finish": "HH:MM",
      "hours": 8.0,
      "total_hours": 8.0,
      "tariff": "rate with unit",
      "in_total": "amount with currency"
    }
  ],
  "job_total_quantity": 8,
  "total_hours": 64.0,
  "job_total_amount": "amount with currency",
  "misc_items": [
    {
      "item": "item name",
      "quantity": "quantity with unit",
      "tariff": "rate with unit",
      "in_total": "amount with currency"
    }
  ],
  "misc_total": "amount with currency",
  "total_price_excl_vat": "amount with currency",
  "vat_percentage": "percentage",
  "vat_amount": "amount with currency",
  "total_price_incl_vat": "amount with currency"
}"""


def extract_invoice_from_bytes(file_bytes: bytes, mime_type: str) -> InvoiceData:
    """Send raw file bytes to Gemini and return structured InvoiceData."""
    client = genai.Client(api_key=GOOGLE_API_KEY)

    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=[
            types.Content(
                role="user",
                parts=[
                    types.Part.from_bytes(data=file_bytes, mime_type=mime_type),
                    types.Part.from_text(text=EXTRACTION_PROMPT),
                ],
            )
        ],
        config=types.GenerateContentConfig(temperature=0.0, max_output_tokens=4096),
    )

    raw = response.text.strip()
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1]
    if raw.endswith("```"):
        raw = raw.rsplit("```", 1)[0]
    raw = raw.strip()

    data = json.loads(raw)
    return InvoiceData(**data)




#  EXCEL EXPORT

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(bold=True, color="2F5496", size=12)
LABEL_FONT = Font(bold=True, size=10)
VALUE_FONT = Font(size=10)
TOTAL_FONT = Font(bold=True, size=11, color="C00000")
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _header_row(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _cell(ws, row, col, val, font=VALUE_FONT, align="left"):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = font
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = THIN_BORDER
    return cell


def export_to_excel(data: InvoiceData, output_path: str) -> str:
    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws.cell(row=row, column=1, value="PRICE QUOTE — INVOICE SUMMARY")
    c.font = Font(bold=True, size=14, color="2F5496")
    c.alignment = Alignment(horizontal="center")
    row += 2

    sections = [
        ("Supplier Information", [
            ("Name", data.supplier_name), ("Address", data.supplier_address),
            ("Registration No.", data.supplier_registration_no),
            ("VAT Number", data.supplier_vat_number),
        ]),
        ("Client Information", [
            ("Name", data.client_name), ("Address", data.client_address),
            ("Registration No.", data.client_registration_no),
            ("VAT Number", data.client_vat_number),
        ]),
        ("Order Details", [
            ("Order Name", data.order_name), ("Order Number", data.order_number),
            ("Date Range", data.date_range), ("Location", data.location),
        ]),
        ("Totals", [
            ("Total Hours", data.total_hours), ("Job Total", data.job_total_amount),
            ("Misc Total", data.misc_total),
            ("Subtotal (excl. VAT)", data.total_price_excl_vat),
            (f"VAT ({data.vat_percentage})", data.vat_amount),
            ("TOTAL (incl. VAT)", data.total_price_incl_vat),
        ]),
    ]

    for title, fields in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        row += 1
        for label, val in fields:
            _cell(ws, row, 1, label, font=LABEL_FONT)
            f = TOTAL_FONT if label == "TOTAL (incl. VAT)" else VALUE_FONT
            c = _cell(ws, row, 2, val, font=f)
            if label == "TOTAL (incl. VAT)":
                c.fill = TOTAL_FILL
            row += 1
        row += 1

    # ── Sheet 2: Job Shifts ──
    ws2 = wb.create_sheet("Job Shifts")
    hdrs = ["Shift", "Date", "Role", "Qty", "Start", "Finish", "Hours", "Total Hours", "Tariff", "Total"]
    widths = [22, 10, 28, 6, 8, 8, 8, 12, 16, 16]
    for i, (h, w) in enumerate(zip(hdrs, widths), 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.cell(row=1, column=i, value=h)
    _header_row(ws2, 1, 1, len(hdrs))

    for r, s in enumerate(data.shifts, 2):
        vals = [s.shift_name, s.date, s.role, s.quantity, s.start, s.finish,
                s.hours, s.total_hours, s.tariff, s.in_total]
        for c, v in enumerate(vals, 1):
            _cell(ws2, r, c, v, align="right" if c >= 4 else "left")

    tr = len(data.shifts) + 2
    _cell(ws2, tr, 1, "TOTAL", font=TOTAL_FONT)
    for c in range(2, len(hdrs) + 1):
        ws2.cell(row=tr, column=c).border = THIN_BORDER
    _cell(ws2, tr, 4, data.job_total_quantity, font=TOTAL_FONT, align="right")
    _cell(ws2, tr, 8, data.total_hours, font=TOTAL_FONT, align="right")
    _cell(ws2, tr, 10, data.job_total_amount, font=TOTAL_FONT, align="right")
    for c in range(1, len(hdrs) + 1):
        ws2.cell(row=tr, column=c).fill = TOTAL_FILL

    # ── Sheet 3: Misc Costs ──
    ws3 = wb.create_sheet("Misc Costs")
    mh = ["Item", "Quantity", "Tariff", "Total"]
    mw = [30, 15, 20, 20]
    for i, (h, w) in enumerate(zip(mh, mw), 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.cell(row=1, column=i, value=h)
    _header_row(ws3, 1, 1, len(mh))

    for r, m in enumerate(data.misc_items, 2):
        for c, v in enumerate([m.item, m.quantity, m.tariff, m.in_total], 1):
            _cell(ws3, r, c, v, align="right" if c >= 2 else "left")

    mr = len(data.misc_items) + 2
    _cell(ws3, mr, 1, "MISC TOTAL", font=TOTAL_FONT)
    for c in range(2, len(mh) + 1):
        ws3.cell(row=mr, column=c).border = THIN_BORDER
    _cell(ws3, mr, 4, data.misc_total, font=TOTAL_FONT, align="right")
    for c in range(1, len(mh) + 1):
        ws3.cell(row=mr, column=c).fill = TOTAL_FILL

    mr += 2
    for label, val in [
        ("Subtotal (excl. VAT)", data.total_price_excl_vat),
        (f"VAT ({data.vat_percentage})", data.vat_amount),
        ("TOTAL (incl. VAT)", data.total_price_incl_vat),
    ]:
        _cell(ws3, mr, 3, label, font=LABEL_FONT, align="right")
        f = TOTAL_FONT if "TOTAL" in label else VALUE_FONT
        c = _cell(ws3, mr, 4, val, font=f, align="right")
        if "TOTAL (incl." in label:
            ws3.cell(row=mr, column=3).fill = TOTAL_FILL
            c.fill = TOTAL_FILL
        mr += 1

    wb.save(output_path)
    return output_path


#  TELEGRAM BOT HANDLERS

SUPPORTED_MIME = {
    "application/pdf": "application/pdf",
    "image/png": "image/png",
    "image/jpeg": "image/jpeg",
    "image/jpg": "image/jpeg",
    "image/webp": "image/webp",
}

WELCOME_MSG = (
    "👋 *Welcome to Invoice OCR Bot!*\n\n"
    "Send me an invoice as a *photo*, *image file*, or *PDF* and I'll extract "
    "the data and return a formatted Excel file.\n\n"
    "📎 *Supported formats:* PNG, JPG, JPEG, WebP, PDF\n"
    "📊 *Output:* .xlsx with Summary, Job Shifts & Misc Costs sheets"
)


async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(WELCOME_MSG, parse_mode="Markdown")


async def cmd_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        " *How to use:*\n\n"
        "1. Send a photo or file of an invoice\n"
        "2. Wait while I extract the data\n"
        "3. Receive your formatted .xlsx file\n\n"
        "That's it!",
        parse_mode="Markdown",
    )


async def handle_photo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle photos sent directly (compressed by Telegram)."""
    msg = await update.message.reply_text("📥 Received photo. Extracting data… ⏳")

    try:
        photo = update.message.photo[-1]  # highest resolution
        file = await ctx.bot.get_file(photo.file_id)
        file_bytes = await file.download_as_bytearray()

        await msg.edit_text(" Running OCR with Gemini 2.5 Flash… ")
        invoice = extract_invoice_from_bytes(bytes(file_bytes), "image/jpeg")

        await msg.edit_text(" Generating Excel file… ⏳")
        xlsx_path = _build_xlsx(invoice, update.effective_user.id)

        await msg.edit_text(" Sending your file…")
        caption = (
            f" *{invoice.order_name or 'Invoice'}*\n"
            f" {invoice.supplier_name} → {invoice.client_name}\n"
            f" Total: *{invoice.total_price_incl_vat}*"
        )
        await update.message.reply_document(
            document=open(xlsx_path, "rb"),
            filename=f"invoice_{invoice.order_number or 'output'}.xlsx",
            caption=caption,
            parse_mode="Markdown",
        )
        os.unlink(xlsx_path)

    except Exception as e:
        logger.error(f"Photo processing error: {e}", exc_info=True)
        await msg.edit_text(f" *Error:* {e}", parse_mode="Markdown")


async def handle_document(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle files sent as documents (PDF, PNG, JPG, etc.)."""
    doc = update.message.document
    mime = doc.mime_type

    if mime not in SUPPORTED_MIME:
        await update.message.reply_text(
            f" Unsupported file type: `{mime}`\n\n"
            "Please send a *PDF, PNG, JPG, or WebP* file.",
            parse_mode="Markdown",
        )
        return

    msg = await update.message.reply_text(
        f" Received `{doc.file_name}`. Extracting data… ⏳",
        parse_mode="Markdown",
    )

    try:
        file = await ctx.bot.get_file(doc.file_id)
        file_bytes = await file.download_as_bytearray()

        await msg.edit_text(" Running OCR with Gemini 2.5 Flash… ⏳")
        invoice = extract_invoice_from_bytes(bytes(file_bytes), SUPPORTED_MIME[mime])

        await msg.edit_text(" Generating Excel file… ⏳")
        xlsx_path = _build_xlsx(invoice, update.effective_user.id)

        await msg.edit_text(" Done! Sending your file…")
        caption = (
            f"*{invoice.order_name or 'Invoice'}*\n"
            f" {invoice.supplier_name} → {invoice.client_name}\n"
            f" Total: *{invoice.total_price_incl_vat}*"
        )
        await update.message.reply_document(
            document=open(xlsx_path, "rb"),
            filename=f"invoice_{invoice.order_number or 'output'}.xlsx",
            caption=caption,
            parse_mode="Markdown",
        )
        os.unlink(xlsx_path)

    except Exception as e:
        logger.error(f"Document processing error: {e}", exc_info=True)
        await msg.edit_text(f" *Error:* {e}", parse_mode="Markdown")


async def handle_unknown(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🤔 Please send me an invoice image or PDF.\nType /help for instructions."
    )


def _build_xlsx(invoice: InvoiceData, user_id: int) -> str:
    """Create a temp xlsx file and return its path."""
    tmp = tempfile.mktemp(
        prefix=f"invoice_{user_id}_",
        suffix=".xlsx",
    )
    export_to_excel(invoice, tmp)
    return tmp


#  MAIN

def main():
    print(" Starting Invoice OCR Bot…")
    app = ApplicationBuilder().token(TELEGRAM_BOT_TOKEN).build()

    # Commands
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_help))

    # Photo (sent as image directly)
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))

    # Document (PDF, image files sent as documents)
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))

    # Catch-all for text or other messages
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_unknown))

    print(" Bot is running. Press Ctrl+C to stop.")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
